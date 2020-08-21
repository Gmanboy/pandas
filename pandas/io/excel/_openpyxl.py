from typing import List

import numpy as np

from pandas._typing import FilePathOrBuffer, Scalar, StorageOptions
from pandas.compat._optional import import_optional_dependency

from pandas.io.excel._base import ExcelWriter, _BaseExcelReader
from pandas.io.excel._util import _validate_freeze_panes


class _OpenpyxlWriter(ExcelWriter):
    engine = "openpyxl"
    supported_extensions = (".xlsx", ".xlsm")

    def __init__(self, path, engine=None, mode="w", **engine_kwargs):
        # Use the openpyxl module as the Excel writer.
        from openpyxl.workbook import Workbook

        super().__init__(path, mode=mode, **engine_kwargs)

        if self.mode == "a":  # Load from existing workbook
            from openpyxl import load_workbook

            book = load_workbook(self.path)
            self.book = book
        else:
            # Create workbook object with default optimized_write=True.
            self.book = Workbook()

            if self.book.worksheets:
                try:
                    self.book.remove(self.book.worksheets[0])
                except AttributeError:

                    # compat - for openpyxl <= 2.4
                    self.book.remove_sheet(self.book.worksheets[0])

    def save(self):
        """
        Save workbook to disk.
        """
        return self.book.save(self.path)

    @classmethod
    def _convert_to_style(cls, style_dict):
        """
        Converts a style_dict to an openpyxl style object.

        Parameters
        ----------
        style_dict : style dictionary to convert
        """
        from openpyxl.style import Style

        xls_style = Style()
        for key, value in style_dict.items():
            for nk, nv in value.items():
                if key == "borders":
                    (
                        xls_style.borders.__getattribute__(nk).__setattr__(
                            "border_style", nv
                        )
                    )
                else:
                    xls_style.__getattribute__(key).__setattr__(nk, nv)

        return xls_style

    @classmethod
    def _convert_to_style_kwargs(cls, style_dict):
        """
        Convert a style_dict to a set of kwargs suitable for initializing
        or updating-on-copy an openpyxl v2 style object.

        Parameters
        ----------
        style_dict : dict
            A dict with zero or more of the following keys (or their synonyms).
                'font'
                'fill'
                'border' ('borders')
                'alignment'
                'number_format'
                'protection'

        Returns
        -------
        style_kwargs : dict
            A dict with the same, normalized keys as ``style_dict`` but each
            value has been replaced with a native openpyxl style object of the
            appropriate class.
        """
        _style_key_map = {"borders": "border"}

        style_kwargs = {}
        for k, v in style_dict.items():
            if k in _style_key_map:
                k = _style_key_map[k]
            _conv_to_x = getattr(cls, f"_convert_to_{k}", lambda x: None)
            new_v = _conv_to_x(v)
            if new_v:
                style_kwargs[k] = new_v

        return style_kwargs

    @classmethod
    def _convert_to_color(cls, color_spec):
        """
        Convert ``color_spec`` to an openpyxl v2 Color object.

        Parameters
        ----------
        color_spec : str, dict
            A 32-bit ARGB hex string, or a dict with zero or more of the
            following keys.
                'rgb'
                'indexed'
                'auto'
                'theme'
                'tint'
                'index'
                'type'

        Returns
        -------
        color : openpyxl.styles.Color
        """
        from openpyxl.styles import Color

        if isinstance(color_spec, str):
            return Color(color_spec)
        else:
            return Color(**color_spec)

    @classmethod
    def _convert_to_font(cls, font_dict):
        """
        Convert ``font_dict`` to an openpyxl v2 Font object.

        Parameters
        ----------
        font_dict : dict
            A dict with zero or more of the following keys (or their synonyms).
                'name'
                'size' ('sz')
                'bold' ('b')
                'italic' ('i')
                'underline' ('u')
                'strikethrough' ('strike')
                'color'
                'vertAlign' ('vertalign')
                'charset'
                'scheme'
                'family'
                'outline'
                'shadow'
                'condense'

        Returns
        -------
        font : openpyxl.styles.Font
        """
        from openpyxl.styles import Font

        _font_key_map = {
            "sz": "size",
            "b": "bold",
            "i": "italic",
            "u": "underline",
            "strike": "strikethrough",
            "vertalign": "vertAlign",
        }

        font_kwargs = {}
        for k, v in font_dict.items():
            if k in _font_key_map:
                k = _font_key_map[k]
            if k == "color":
                v = cls._convert_to_color(v)
            font_kwargs[k] = v

        return Font(**font_kwargs)

    @classmethod
    def _convert_to_stop(cls, stop_seq):
        """
        Convert ``stop_seq`` to a list of openpyxl v2 Color objects,
        suitable for initializing the ``GradientFill`` ``stop`` parameter.

        Parameters
        ----------
        stop_seq : iterable
            An iterable that yields objects suitable for consumption by
            ``_convert_to_color``.

        Returns
        -------
        stop : list of openpyxl.styles.Color
        """
        return map(cls._convert_to_color, stop_seq)

    @classmethod
    def _convert_to_fill(cls, fill_dict):
        """
        Convert ``fill_dict`` to an openpyxl v2 Fill object.

        Parameters
        ----------
        fill_dict : dict
            A dict with one or more of the following keys (or their synonyms),
                'fill_type' ('patternType', 'patterntype')
                'start_color' ('fgColor', 'fgcolor')
                'end_color' ('bgColor', 'bgcolor')
            or one or more of the following keys (or their synonyms).
                'type' ('fill_type')
                'degree'
                'left'
                'right'
                'top'
                'bottom'
                'stop'

        Returns
        -------
        fill : openpyxl.styles.Fill
        """
        from openpyxl.styles import GradientFill, PatternFill

        _pattern_fill_key_map = {
            "patternType": "fill_type",
            "patterntype": "fill_type",
            "fgColor": "start_color",
            "fgcolor": "start_color",
            "bgColor": "end_color",
            "bgcolor": "end_color",
        }

        _gradient_fill_key_map = {"fill_type": "type"}

        pfill_kwargs = {}
        gfill_kwargs = {}
        for k, v in fill_dict.items():
            pk = gk = None
            if k in _pattern_fill_key_map:
                pk = _pattern_fill_key_map[k]
            if k in _gradient_fill_key_map:
                gk = _gradient_fill_key_map[k]
            if pk in ["start_color", "end_color"]:
                v = cls._convert_to_color(v)
            if gk == "stop":
                v = cls._convert_to_stop(v)
            if pk:
                pfill_kwargs[pk] = v
            elif gk:
                gfill_kwargs[gk] = v
            else:
                pfill_kwargs[k] = v
                gfill_kwargs[k] = v

        try:
            return PatternFill(**pfill_kwargs)
        except TypeError:
            return GradientFill(**gfill_kwargs)

    @classmethod
    def _convert_to_side(cls, side_spec):
        """
        Convert ``side_spec`` to an openpyxl v2 Side object.

        Parameters
        ----------
        side_spec : str, dict
            A string specifying the border style, or a dict with zero or more
            of the following keys (or their synonyms).
                'style' ('border_style')
                'color'

        Returns
        -------
        side : openpyxl.styles.Side
        """
        from openpyxl.styles import Side

        _side_key_map = {"border_style": "style"}

        if isinstance(side_spec, str):
            return Side(style=side_spec)

        side_kwargs = {}
        for k, v in side_spec.items():
            if k in _side_key_map:
                k = _side_key_map[k]
            if k == "color":
                v = cls._convert_to_color(v)
            side_kwargs[k] = v

        return Side(**side_kwargs)

    @classmethod
    def _convert_to_border(cls, border_dict):
        """
        Convert ``border_dict`` to an openpyxl v2 Border object.

        Parameters
        ----------
        border_dict : dict
            A dict with zero or more of the following keys (or their synonyms).
                'left'
                'right'
                'top'
                'bottom'
                'diagonal'
                'diagonal_direction'
                'vertical'
                'horizontal'
                'diagonalUp' ('diagonalup')
                'diagonalDown' ('diagonaldown')
                'outline'

        Returns
        -------
        border : openpyxl.styles.Border
        """
        from openpyxl.styles import Border

        _border_key_map = {"diagonalup": "diagonalUp", "diagonaldown": "diagonalDown"}

        border_kwargs = {}
        for k, v in border_dict.items():
            if k in _border_key_map:
                k = _border_key_map[k]
            if k == "color":
                v = cls._convert_to_color(v)
            if k in ["left", "right", "top", "bottom", "diagonal"]:
                v = cls._convert_to_side(v)
            border_kwargs[k] = v

        return Border(**border_kwargs)

    @classmethod
    def _convert_to_alignment(cls, alignment_dict):
        """
        Convert ``alignment_dict`` to an openpyxl v2 Alignment object.

        Parameters
        ----------
        alignment_dict : dict
            A dict with zero or more of the following keys (or their synonyms).
                'horizontal'
                'vertical'
                'text_rotation'
                'wrap_text'
                'shrink_to_fit'
                'indent'
        Returns
        -------
        alignment : openpyxl.styles.Alignment
        """
        from openpyxl.styles import Alignment

        return Alignment(**alignment_dict)

    @classmethod
    def _convert_to_number_format(cls, number_format_dict):
        """
        Convert ``number_format_dict`` to an openpyxl v2.1.0 number format
        initializer.

        Parameters
        ----------
        number_format_dict : dict
            A dict with zero or more of the following keys.
                'format_code' : str

        Returns
        -------
        number_format : str
        """
        return number_format_dict["format_code"]

    @classmethod
    def _convert_to_protection(cls, protection_dict):
        """
        Convert ``protection_dict`` to an openpyxl v2 Protection object.

        Parameters
        ----------
        protection_dict : dict
            A dict with zero or more of the following keys.
                'locked'
                'hidden'

        Returns
        -------
        """
        from openpyxl.styles import Protection

        return Protection(**protection_dict)

    def write_cells(
        self, cells, sheet_name=None, startrow=0, startcol=0, freeze_panes=None
    ):
        # Write the frame cells using openpyxl.
        sheet_name = self._get_sheet_name(sheet_name)

        _style_cache = {}

        if sheet_name in self.sheets:
            wks = self.sheets[sheet_name]
        else:
            wks = self.book.create_sheet()
            wks.title = sheet_name
            self.sheets[sheet_name] = wks

        if _validate_freeze_panes(freeze_panes):
            wks.freeze_panes = wks.cell(
                row=freeze_panes[0] + 1, column=freeze_panes[1] + 1
            )

        for cell in cells:
            xcell = wks.cell(
                row=startrow + cell.row + 1, column=startcol + cell.col + 1
            )
            xcell.value, fmt = self._value_with_fmt(cell.val)
            if fmt:
                xcell.number_format = fmt

            style_kwargs = {}
            if cell.style:
                key = str(cell.style)
                style_kwargs = _style_cache.get(key)
                if style_kwargs is None:
                    style_kwargs = self._convert_to_style_kwargs(cell.style)
                    _style_cache[key] = style_kwargs

            if style_kwargs:
                for k, v in style_kwargs.items():
                    setattr(xcell, k, v)

            if cell.mergestart is not None and cell.mergeend is not None:

                wks.merge_cells(
                    start_row=startrow + cell.row + 1,
                    start_column=startcol + cell.col + 1,
                    end_column=startcol + cell.mergeend + 1,
                    end_row=startrow + cell.mergestart + 1,
                )

                # When cells are merged only the top-left cell is preserved
                # The behaviour of the other cells in a merged range is
                # undefined
                if style_kwargs:
                    first_row = startrow + cell.row + 1
                    last_row = startrow + cell.mergestart + 1
                    first_col = startcol + cell.col + 1
                    last_col = startcol + cell.mergeend + 1

                    for row in range(first_row, last_row + 1):
                        for col in range(first_col, last_col + 1):
                            if row == first_row and col == first_col:
                                # Ignore first cell. It is already handled.
                                continue
                            xcell = wks.cell(column=col, row=row)
                            for k, v in style_kwargs.items():
                                setattr(xcell, k, v)


class _OpenpyxlReader(_BaseExcelReader):
    def __init__(
        self,
        filepath_or_buffer: FilePathOrBuffer,
        storage_options: StorageOptions = None,
    ) -> None:
        """
        Reader using openpyxl engine.

        Parameters
        ----------
        filepath_or_buffer : string, path object or Workbook
            Object to be parsed.
        storage_options : StorageOptions
            passed to fsspec for appropriate URLs (see ``get_filepath_or_buffer``)
        """
        import_optional_dependency("openpyxl")
        super().__init__(filepath_or_buffer, storage_options=storage_options)

    @property
    def _workbook_class(self):
        from openpyxl import Workbook

        return Workbook

    def load_workbook(self, filepath_or_buffer: FilePathOrBuffer):
        from openpyxl import load_workbook

        return load_workbook(
            filepath_or_buffer, read_only=True, data_only=True, keep_links=False
        )

    def close(self):
        # https://stackoverflow.com/questions/31416842/
        #  openpyxl-does-not-close-excel-workbook-in-read-only-mode
        self.book.close()

    @property
    def sheet_names(self) -> List[str]:
        return self.book.sheetnames

    def get_sheet_by_name(self, name: str):
        return self.book[name]

    def get_sheet_by_index(self, index: int):
        return self.book.worksheets[index]

    def _convert_cell(self, cell, convert_float: bool) -> Scalar:

        # TODO: replace with openpyxl constants
        if cell.is_date:
            return cell.value
        elif cell.data_type == "e":
            return np.nan
        elif cell.data_type == "b":
            return bool(cell.value)
        elif cell.value is None:
            return ""  # compat with xlrd
        elif cell.data_type == "n":
            # GH5394
            if convert_float:
                val = int(cell.value)
                if val == cell.value:
                    return val
            else:
                return float(cell.value)

        return cell.value

    def get_sheet_data(self, sheet, convert_float: bool) -> List[List[Scalar]]:
        data: List[List[Scalar]] = []
        for row in sheet.rows:
            data.append([self._convert_cell(cell, convert_float) for cell in row])

        return data
